from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import xlrd

from .excel_import import import_order, normalize_header, parse_decimal


ORDER_FILE_PATTERNS = ("*.xls", "*.xlsx", "*.xlsm")


@dataclass(frozen=True)
class OrderFile:
    path: Path
    modified_at: float
    order_number: str = ""
    order_date: str = ""
    customer_company: str = ""
    customer_name: str = ""
    customer_city: str = ""
    customer_zip: str = ""
    customer_phone: str = ""
    customer_email: str = ""
    customer_vat: str = ""
    shipping_method: str = ""
    line_count: int = 0
    package_count: int = 0
    piece_count: int = 0
    total_amount: Decimal | None = None
    error: str = ""

    @property
    def display_name(self) -> str:
        customer = self.customer_name or self.customer_company or "Client inconnu"
        amount = f"{self.total_amount} EUR" if self.total_amount is not None else "total inconnu"
        return f"{self.order_number or self.path.stem} - {customer} - {amount}"

    @property
    def modified_at_label(self) -> str:
        return datetime.fromtimestamp(self.modified_at).strftime("%d/%m/%Y %H:%M")


def list_order_files(folder: str | Path) -> list[OrderFile]:
    folder_path = Path(folder).expanduser()
    if not folder_path.exists() or not folder_path.is_dir():
        return []

    files: list[OrderFile] = []
    for pattern in ORDER_FILE_PATTERNS:
        for path in folder_path.glob(pattern):
            if not path.is_file() or path.name.startswith("~$"):
                continue
            if not path.stem.isdigit():
                continue
            files.append(summarize_order_file(path))

    return sorted(files, key=lambda item: (item.modified_at, item.path.name), reverse=True)


def latest_order_file(folder: str | Path) -> Path | None:
    files = list_order_files(folder)
    return files[0].path if files else None


def summarize_order_file(path: str | Path) -> OrderFile:
    file_path = Path(path)
    modified_at = file_path.stat().st_mtime
    try:
        metadata = _read_microstore_metadata(file_path)
        order_rows = import_order(file_path).rows
        package_count = sum(row.package_count for row in order_rows)
        piece_count = sum(row.quantity_pieces or 0 for row in order_rows)
        total_amount = sum(
            (row.unit_price_ht or Decimal("0")) * (row.quantity_pieces or 0)
            for row in order_rows
        )
        return OrderFile(
            path=file_path,
            modified_at=modified_at,
            order_number=str(metadata.get("order_sn") or file_path.stem).strip(),
            order_date=str(metadata.get("date") or "").strip(),
            customer_company=str(metadata.get("customer_company") or "").strip(),
            customer_name=str(metadata.get("customer_name") or "").strip(),
            customer_city=str(metadata.get("customer_city") or "").strip(),
            customer_zip=str(metadata.get("customer_zip") or "").strip(),
            customer_phone=str(metadata.get("customer_phone") or "").strip(),
            customer_email=str(metadata.get("customer_mail") or "").strip(),
            customer_vat=str(metadata.get("customer_company_id") or "").strip(),
            shipping_method=str(metadata.get("shipping_method") or "").strip(),
            line_count=len(order_rows),
            package_count=package_count,
            piece_count=piece_count,
            total_amount=total_amount,
        )
    except Exception as exc:
        return OrderFile(path=file_path, modified_at=modified_at, order_number=file_path.stem, error=str(exc))


def _read_microstore_metadata(path: Path) -> dict[str, Any]:
    if path.suffix.lower() == ".xls":
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows < 2:
            return {}
        keys = [normalize_header(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        values = [sheet.cell_value(1, col) for col in range(sheet.ncols)]
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
        finally:
            workbook.close()
        if len(rows) < 2:
            return {}
        keys = [normalize_header(value) for value in rows[0]]
        values = list(rows[1])

    metadata: dict[str, Any] = {}
    for key, value in zip(keys, values):
        if key:
            metadata[key.replace(" ", "_")] = _clean_metadata_value(value)
    return metadata


def _clean_metadata_value(value: Any) -> Any:
    if value in (None, ""):
        return ""
    decimal_value = parse_decimal(value)
    if decimal_value is not None and not isinstance(value, str):
        return str(value).rstrip("0").rstrip(".")
    return str(value).strip()
