from __future__ import annotations

import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import xlrd

from .models import Product


PRODUCT_HEADER_ALIASES = {
    "ref": {
        "reference",
        "reference produit",
        "ref",
        "sku",
        "code",
        "code article",
        "modele",
        "modele article",
        "article no",
        "article",
    },
    "type_label": {
        "type",
        "categorie",
        "category",
        "classification",
        "famille",
        "type article",
        "categorie article",
    },
    "unit_price_ht": {
        "prix",
        "prix ht",
        "pu ht",
        "p.u. ht",
        "prix unitaire",
        "unit price",
    },
    "package_size": {
        "colisage",
        "pcs/ctn",
        "pieces par paquet",
        "piece par paquet",
        "qte paquet",
        "pack",
        "package",
    },
}

ORDER_HEADER_ALIASES = {
    "ref": PRODUCT_HEADER_ALIASES["ref"]
    | {
        "n de produits",
        "no de produits",
        "numero de produits",
        "n produits",
        "produit",
    },
    "package_count": {
        "quantite",
        "qte",
        "qty",
        "quantity",
        "nombre",
    },
    "package_size": PRODUCT_HEADER_ALIASES["package_size"],
    "quantity_pieces": {
        "nombre de pieces",
        "nombre piece",
        "quantite pieces",
        "qte pieces",
        "pieces",
        "nombre de pieces quantite unites de colisage",
        "nombre de pieces quantite unites de colisage",
    },
    "unit_price_ht": PRODUCT_HEADER_ALIASES["unit_price_ht"],
}

MICROSTORE_XLS_ORDER_HEADERS = {
    "product reference",
    "quantity",
    "unit",
    "unit price",
}


@dataclass(frozen=True)
class OrderRow:
    ref: str
    package_count: int
    package_size: int | None = None
    quantity_pieces: int | None = None
    unit_price_ht: Decimal | None = None


@dataclass(frozen=True)
class ImportResult:
    rows: list[Product] | list[OrderRow]
    warnings: list[str]


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    for old in ("_", "-", ".", "°", "(", ")", "*"):
        text = text.replace(old, " ")
    return " ".join(text.split())


def normalize_ref(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\u00a0", " ").replace("\u200e", "").replace("\u200f", "")
    return "".join(text.split()).upper()


def parse_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    text = str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(Decimal(str(value).strip().replace(",", ".")))
    except (InvalidOperation, ValueError):
        return None


def _load_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if path.suffix.lower() == ".xls":
        return _load_xls_rows(path)
    return _load_xlsx_rows(path)


def _load_xlsx_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not raw_rows:
        return [], []

    header_index = _detect_header_row(raw_rows)
    headers = [normalize_header(cell) for cell in raw_rows[header_index]]
    rows: list[dict[str, Any]] = []
    for raw in raw_rows[header_index + 1 :]:
        if not any(cell not in (None, "") for cell in raw):
            continue
        rows.append({headers[index]: value for index, value in enumerate(raw) if index < len(headers)})
    return headers, rows


def _load_xls_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    raw_rows: list[tuple[Any, ...]] = [
        tuple(sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols))
        for row_index in range(sheet.nrows)
    ]
    if not raw_rows:
        return [], []

    header_index = _detect_header_row(raw_rows)
    headers = [normalize_header(cell) for cell in raw_rows[header_index]]
    rows: list[dict[str, Any]] = []
    for raw in raw_rows[header_index + 1 :]:
        if not any(cell not in (None, "") for cell in raw):
            continue
        rows.append({headers[index]: value for index, value in enumerate(raw) if index < len(headers)})
    return headers, rows


def _detect_header_row(raw_rows: list[tuple[Any, ...]]) -> int:
    alias_values: set[str] = set()
    for names in PRODUCT_HEADER_ALIASES.values():
        alias_values.update(normalize_header(name) for name in names)
    for names in ORDER_HEADER_ALIASES.values():
        alias_values.update(normalize_header(name) for name in names)

    best_index = 0
    best_score = -1
    for index, row in enumerate(raw_rows[:10]):
        headers = {normalize_header(cell) for cell in row if cell not in (None, "")}
        score = len(headers & alias_values)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def _match_columns(headers: list[str], aliases: dict[str, set[str]]) -> dict[str, str]:
    matched: dict[str, str] = {}
    for field, names in aliases.items():
        normalized_aliases = {normalize_header(name) for name in names}
        for header in headers:
            if header in normalized_aliases:
                matched[field] = header
                break
    return matched


def import_products(path: Path) -> ImportResult:
    headers, rows = _load_rows(path)
    columns = _match_columns(headers, PRODUCT_HEADER_ALIASES)
    warnings: list[str] = []
    if "ref" not in columns:
        raise ValueError("Colonne reference introuvable dans l'export produits.")

    products: list[Product] = []
    for index, row in enumerate(rows, start=2):
        ref = normalize_ref(row.get(columns["ref"], ""))
        if not ref:
            warnings.append(f"Ligne {index}: reference vide ignoree")
            continue
        product = Product(
            id=None,
            ref=ref,
            type_label=str(row.get(columns.get("type_label", ""), "") or "").strip(),
            name="",
            unit_price_ht=parse_decimal(row.get(columns.get("unit_price_ht", ""))),
            package_size=parse_int(row.get(columns.get("package_size", ""))),
        )
        products.append(product)
    return ImportResult(rows=products, warnings=warnings)


def import_order(path: Path) -> ImportResult:
    headers, rows = _load_rows(path)
    if _is_microstore_order_xls(headers):
        return _import_microstore_order_xls_rows(rows)

    columns = _match_columns(headers, ORDER_HEADER_ALIASES)
    warnings: list[str] = []
    if "ref" not in columns:
        raise ValueError("Colonne reference introuvable dans l'export commande.")
    if "package_count" not in columns and "quantity_pieces" not in columns:
        raise ValueError("Colonne quantite introuvable dans l'export commande.")

    order_rows: list[OrderRow] = []
    for index, row in enumerate(rows, start=2):
        ref = normalize_ref(row.get(columns["ref"], ""))
        package_count = parse_int(row.get(columns.get("package_count", ""))) or 0
        package_size = parse_int(row.get(columns.get("package_size", "")))
        quantity_pieces = parse_int(row.get(columns.get("quantity_pieces", "")))
        if not ref:
            warnings.append(f"Ligne {index}: reference vide ignoree")
            continue
        if package_count <= 0 and (not quantity_pieces or quantity_pieces <= 0):
            warnings.append(f"Ligne {index}: quantite invalide pour {ref}")
            continue
        if quantity_pieces and package_count and package_size and quantity_pieces != package_count * package_size:
            warnings.append(
                f"Ligne {index}: controle quantite different pour {ref} "
                f"({package_count} x {package_size} != {quantity_pieces})"
            )
        order_rows.append(
            OrderRow(
                ref=ref,
                package_count=package_count,
                package_size=package_size,
                quantity_pieces=quantity_pieces,
                unit_price_ht=parse_decimal(row.get(columns.get("unit_price_ht", ""))),
            )
        )
    return ImportResult(rows=order_rows, warnings=warnings)


def _is_microstore_order_xls(headers: list[str]) -> bool:
    return MICROSTORE_XLS_ORDER_HEADERS.issubset(set(headers))


def _import_microstore_order_xls_rows(rows: list[dict[str, Any]]) -> ImportResult:
    warnings: list[str] = []
    order_rows: list[OrderRow] = []
    for index, row in enumerate(rows, start=1):
        ref = normalize_ref(row.get("product reference", ""))
        if not ref or ref == "TOTAL:":
            continue

        package_count = parse_int(row.get("quantity")) or 0
        package_size = parse_int(row.get("unit"))
        quantity_pieces = package_count * package_size if package_count > 0 and package_size else None
        if package_count <= 0:
            warnings.append(f"Ligne commande {index}: quantite invalide pour {ref}")
            continue
        if not package_size:
            warnings.append(f"Ligne commande {index}: colisage absent pour {ref}")

        order_rows.append(
            OrderRow(
                ref=ref,
                package_count=package_count,
                package_size=package_size,
                quantity_pieces=quantity_pieces,
                unit_price_ht=parse_decimal(row.get("unit price")),
            )
        )
    return ImportResult(rows=order_rows, warnings=warnings)
