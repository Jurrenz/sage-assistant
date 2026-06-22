from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from app.excel_import import normalize_ref
from app.models import utc_now_iso
from telegram_bot.database.warehouse_repository import WarehouseRepository, WarehouseStockRecord, WarehouseImportSummary


SHEET_NAME = "STOCK"
REF_COLUMN = 9
TAIL_PIECES_COLUMN = 41
PIECES_PER_BOX_COLUMN = 42
BOX_COUNT_COLUMN = 43


def parse_stock_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    matches = re.findall(r"-?\d+(?:[\.,]\d+)?", text)
    if matches:
        total = Decimal("0")
        for match in matches:
            try:
                total += Decimal(match.replace(",", "."))
            except InvalidOperation:
                return None
        return int(total)
    try:
        number = Decimal(text.replace(",", "."))
    except InvalidOperation:
        return None
    return int(number)


def format_stock_display(tail_pieces: int | None, pieces_per_box: int | None, box_count: int | None) -> str:
    tail = tail_pieces or 0
    pieces = pieces_per_box or 0
    boxes = box_count or 0
    parts: list[str] = []
    if tail > 0:
        parts.append(f"({tail}p)")
    if pieces > 0 and boxes > 0:
        parts.append(f"{pieces}p×{boxes}")
    return "+".join(parts) if parts else "0"


def calculate_total_pieces(tail_pieces: int | None, pieces_per_box: int | None, box_count: int | None) -> int:
    return int(tail_pieces or 0) + int(pieces_per_box or 0) * int(box_count or 0)


def calculate_total_packages(total_pieces: int, package_size: int | None) -> int | None:
    if not package_size or package_size <= 0:
        return None
    return total_pieces // package_size


class StockImportService:
    def __init__(self, repository: WarehouseRepository) -> None:
        self.repository = repository

    def sync_from_xlsx(self, path: Path) -> WarehouseImportSummary:
        if not path.exists():
            raise FileNotFoundError(f"Fichier stock introuvable: {path}")
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook[SHEET_NAME.lower()]
            records: list[WarehouseStockRecord] = []
            ignored_refs: list[str] = []
            synced_at = utc_now_iso()
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                ref = normalize_ref(row[REF_COLUMN - 1] if len(row) >= REF_COLUMN else "")
                if not ref:
                    continue
                if not self.repository.product_exists(ref):
                    ignored_refs.append(ref)
                    continue
                tail_pieces = parse_stock_int(row[TAIL_PIECES_COLUMN - 1] if len(row) >= TAIL_PIECES_COLUMN else None)
                pieces_per_box = parse_stock_int(row[PIECES_PER_BOX_COLUMN - 1] if len(row) >= PIECES_PER_BOX_COLUMN else None)
                box_count = parse_stock_int(row[BOX_COUNT_COLUMN - 1] if len(row) >= BOX_COUNT_COLUMN else None)
                total_pieces = calculate_total_pieces(tail_pieces, pieces_per_box, box_count)
                package_size = self.repository.package_size_for_ref(ref)
                records.append(
                    WarehouseStockRecord(
                        ref=ref,
                        tail_pieces=tail_pieces,
                        pieces_per_box=pieces_per_box,
                        box_count=box_count,
                        display_text=format_stock_display(tail_pieces, pieces_per_box, box_count),
                        total_pieces=total_pieces,
                        total_packages=calculate_total_packages(total_pieces, package_size),
                        source_row=row_index,
                        last_synced_at=synced_at,
                    )
                )
        finally:
            workbook.close()
        return self.repository.replace_from_import(records, ignored_refs)
