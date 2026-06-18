from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.customs import (
    MAX_PARCEL_WEIGHT_KG,
    CustomsDeclarationDraft,
    CustomsLine,
    CustomsParcel,
    add_quantity_to_parcel,
    adapt_parcel_weights,
    build_customs_draft,
    build_laposte_script,
    customs_line_from_invoice_line,
    draft_from_payload,
    draft_to_payload,
    export_packing_list_xlsx,
    packing_list_text,
    remaining_quantities_by_ref,
    set_parcel_line_quantity,
    set_draft_parcel_count,
    parcel_to_laposte_payload,
)
from app.db import Database
from app.models import InvoiceLine, Product


def test_customs_line_uses_product_weight_and_invoice_value():
    line = InvoiceLine(
        ref="FL395-2",
        sage_code="RO",
        description="ROBE / TUNIC FL395-2",
        quantity_pieces=12,
        unit_price_ht=Decimal("5.50"),
    )
    product = Product(
        id=1,
        ref="FL395-2",
        type_label="ROBES COURTES",
        name="Robe",
        unit_price_ht=Decimal("5.50"),
        package_size=12,
        weight_grams=310,
        origin_country="China",
    )

    customs_line = customs_line_from_invoice_line(line, product)

    assert customs_line.unit_weight_kg == Decimal("0.310")
    assert customs_line.total_weight_kg == Decimal("3.720")
    assert customs_line.total_value_ht == Decimal("66.00")
    assert customs_line.origin_country == "China"
    assert customs_line.hs_number == "62044300"


def test_customs_line_falls_back_to_sage_code_weight():
    line = InvoiceLine(ref="PANT-1", sage_code="PA", description="Pantalon", quantity_pieces=10, unit_price_ht=Decimal("4.20"))

    customs_line = customs_line_from_invoice_line(line, None)

    assert customs_line.unit_weight_kg == Decimal("0.200")
    assert customs_line.total_weight_kg == Decimal("2.000")
    assert customs_line.total_value_ht == Decimal("42.00")


def test_customs_line_uses_heavier_defaults_for_ensembles_and_combis():
    ensemble = InvoiceLine(ref="EN-1", sage_code="EN", description="Ensemble", quantity_pieces=4, unit_price_ht=Decimal("10.00"))
    combi = InvoiceLine(ref="CO-1", sage_code="CO", description="Combi", quantity_pieces=3, unit_price_ht=Decimal("12.00"))

    ensemble_line = customs_line_from_invoice_line(ensemble, None)
    combi_line = customs_line_from_invoice_line(combi, None)

    assert ensemble_line.unit_weight_kg == Decimal("0.500")
    assert ensemble_line.total_weight_kg == Decimal("2.000")
    assert combi_line.unit_weight_kg == Decimal("0.400")
    assert combi_line.total_weight_kg == Decimal("1.200")


def test_customs_line_keeps_package_size_in_payload():
    line = InvoiceLine(
        ref="FL395-2",
        sage_code="RO",
        description="Robe",
        quantity_pieces=24,
        package_count=2,
        package_size=12,
        unit_price_ht=Decimal("5.50"),
    )

    draft = build_customs_draft("Microstore", "1001", [line], {}, {})
    payload = draft_to_payload(draft)
    restored = draft_from_payload(payload)

    assert draft.parcels[0].lines[0].package_size == 12
    assert payload["parcels"][0]["lines"][0]["package_size"] == 12
    assert restored is not None
    assert restored.parcels[0].lines[0].package_size == 12


def test_customs_legacy_payload_without_package_size_still_loads():
    payload = {
        "source": "Microstore",
        "order_key": "1001",
        "parcels": [
            {
                "name": "Colis 1",
                "lines": [
                    {
                        "ref": "FL395-2",
                        "description": "Robe",
                        "sage_code": "RO",
                        "quantity": 12,
                        "unit_weight_kg": "0.300",
                        "unit_value_ht": "5.50",
                    }
                ],
            }
        ],
    }

    restored = draft_from_payload(payload)

    assert restored is not None
    assert restored.parcels[0].lines[0].package_size is None


def test_parcel_reports_weight_over_30kg_and_over_real_limit():
    line = InvoiceLine(ref="FL395-2", sage_code="RO", description="Robe", quantity_pieces=101, unit_price_ht=Decimal("5.50"))
    customs_line = customs_line_from_invoice_line(line, None)
    parcel = CustomsParcel(max_weight_kg=MAX_PARCEL_WEIGHT_KG, lines=[customs_line])

    assert parcel.total_weight_kg == Decimal("30.300")
    assert parcel.weight_errors() == ["Colis 1: poids declare 30.300 kg > plafond 30.00 kg"]

    parcel.max_weight_kg = Decimal("8.00")
    assert parcel.weight_errors() == ["Colis 1: poids declare 30.300 kg > plafond 8.00 kg"]


def test_adapt_parcel_weights_targets_parcel_max_weight():
    parcel = CustomsParcel(
        max_weight_kg=Decimal("8.00"),
        lines=[
            CustomsLine("PA-1", "Pantalon", "PA", 10, Decimal("0.200"), Decimal("4.20")),
            CustomsLine("ZERO", "Zero", "RO", 0, Decimal("0.300"), Decimal("5.50")),
        ],
    )

    adapt_parcel_weights(parcel, parcel.max_weight_kg)

    assert parcel.lines[0].unit_weight_kg == Decimal("0.799")
    assert parcel.lines[1].unit_weight_kg == Decimal("0.300")
    assert parcel.total_weight_kg == Decimal("7.990")
    assert parcel.total_weight_kg <= parcel.max_weight_kg


def test_adapt_parcel_weights_never_exceeds_30kg_after_rounding():
    parcel = CustomsParcel(
        max_weight_kg=Decimal("30.00"),
        lines=[
            CustomsLine("A", "Article A", "RO", 17, Decimal("0.731"), Decimal("5.00")),
            CustomsLine("B", "Article B", "EN", 23, Decimal("0.482"), Decimal("6.00")),
            CustomsLine("C", "Article C", "PA", 31, Decimal("0.211"), Decimal("4.00")),
        ],
    )

    adapt_parcel_weights(parcel, parcel.max_weight_kg)

    assert parcel.total_weight_kg <= Decimal("30.000")
    assert parcel.total_weight_kg <= parcel.max_weight_kg


def test_customs_draft_remove_parcel_renumbers_and_keeps_last_parcel():
    draft = CustomsDeclarationDraft(
        source="Microstore",
        order_key="1001",
        parcels=[
            CustomsParcel(name="Colis 1"),
            CustomsParcel(name="Colis 2"),
            CustomsParcel(name="Colis 3"),
        ],
    )

    assert draft.remove_parcel(1) is True
    assert [parcel.name for parcel in draft.parcels] == ["Colis 1", "Colis 2"]
    assert draft.remove_parcel(0) is True
    assert [parcel.name for parcel in draft.parcels] == ["Colis 1"]
    assert draft.remove_parcel(0) is False
    assert [parcel.name for parcel in draft.parcels] == ["Colis 1"]


def test_packing_helpers_cap_added_quantities_and_track_remaining():
    source = CustomsDeclarationDraft(
        source="Microstore",
        order_key="1001",
        parcels=[
            CustomsParcel(
                name="Colis 1",
                lines=[
                    CustomsLine("FL395-2", "Robe", "RO", 24, Decimal("0.300"), Decimal("5.50"), package_size=12),
                    CustomsLine("PA-1", "Pantalon", "PA", 5, Decimal("0.200"), Decimal("4.20"), package_size=5),
                ],
            )
        ],
    )
    draft = CustomsDeclarationDraft(source="Microstore", order_key="1001", parcels=[CustomsParcel(name="Colis 1")])
    set_draft_parcel_count(draft, source, 2)

    added_pack = add_quantity_to_parcel(draft, source, 0, "FL395-2", source.parcels[0].lines[0].package_size or 1)
    added_overflow = add_quantity_to_parcel(draft, source, 1, "FL395-2", 99)

    assert added_pack == 12
    assert added_overflow == 12
    assert remaining_quantities_by_ref(draft, source)["FL395-2"] == 0
    assert draft.parcels[0].lines[0].quantity == 12
    assert draft.parcels[1].lines[0].quantity == 12


def test_packing_helpers_rebuild_and_renumber_parcels():
    source = CustomsDeclarationDraft(
        source="Microstore",
        order_key="1001",
        parcels=[
            CustomsParcel(
                name="Colis 1",
                lines=[CustomsLine("A", "Article A", "RO", 10, Decimal("0.300"), Decimal("5.00"))],
            )
        ],
    )
    draft = CustomsDeclarationDraft(source="Microstore", order_key="1001", parcels=[CustomsParcel(name="Old")])

    set_draft_parcel_count(draft, source, 3)
    add_quantity_to_parcel(draft, source, 2, "A", 4)

    assert [parcel.name for parcel in draft.parcels] == ["Colis 1", "Colis 2", "Colis 3"]
    assert all(parcel.lines[0].ref == "A" for parcel in draft.parcels)
    assert draft.parcels[2].lines[0].quantity == 4


def test_set_parcel_line_quantity_sets_final_quantity_and_caps_to_available():
    source = CustomsDeclarationDraft(
        source="Microstore",
        order_key="1001",
        parcels=[CustomsParcel(name="Colis 1", lines=[CustomsLine("A", "Article A", "RO", 10, Decimal("0.300"), Decimal("5.00"))])],
    )
    draft = CustomsDeclarationDraft(
        source="Microstore",
        order_key="1001",
        parcels=[
            CustomsParcel(name="Colis 1", lines=[CustomsLine("A", "Article A", "RO", 3, Decimal("0.300"), Decimal("5.00"))]),
            CustomsParcel(name="Colis 2", lines=[CustomsLine("A", "Article A", "RO", 4, Decimal("0.300"), Decimal("5.00"))]),
        ],
    )

    assert set_parcel_line_quantity(draft, source, 0, "A", 6) == 6
    assert draft.parcels[0].lines[0].quantity == 6
    assert set_parcel_line_quantity(draft, source, 0, "A", 99) == 6
    assert draft.parcels[0].lines[0].quantity == 6


def test_set_parcel_line_quantity_empty_or_negative_target_becomes_zero():
    source = CustomsDeclarationDraft(
        source="Microstore",
        order_key="1001",
        parcels=[CustomsParcel(name="Colis 1", lines=[CustomsLine("A", "Article A", "RO", 10, Decimal("0.300"), Decimal("5.00"))])],
    )
    draft = CustomsDeclarationDraft(
        source="Microstore",
        order_key="1001",
        parcels=[CustomsParcel(name="Colis 1", lines=[CustomsLine("A", "Article A", "RO", 3, Decimal("0.300"), Decimal("5.00"))])],
    )

    assert set_parcel_line_quantity(draft, source, 0, "A", -5) == 0
    assert draft.parcels[0].lines[0].quantity == 0


def test_packing_list_text_ignores_zero_quantity_lines():
    draft = CustomsDeclarationDraft(
        source="Microstore",
        order_key="1001",
        parcels=[
            CustomsParcel(
                name="Colis 1",
                lines=[
                    CustomsLine("LA57-5", "Robe", "RO", 12, Decimal("0.300"), Decimal("5.50")),
                    CustomsLine("ZERO", "Zero", "PA", 0, Decimal("0.200"), Decimal("4.00")),
                ],
            ),
            CustomsParcel(
                name="Colis 2",
                lines=[CustomsLine("G05-3", "Top", "SH", 6, Decimal("0.200"), Decimal("4.20"))],
            ),
        ],
    )

    text = packing_list_text(draft)

    assert text == "Colis 1: LA57-5 x12\nColis 2: G05-3 x6"
    assert "ZERO" not in text


def test_export_packing_list_xlsx_contains_summary_lines_and_totals(tmp_path):
    draft = CustomsDeclarationDraft(
        source="Microstore",
        order_key="1001",
        parcels=[
            CustomsParcel(
                name="Colis 1",
                lines=[
                    CustomsLine("LA57-5", "Robe", "RO", 12, Decimal("0.300"), Decimal("5.50")),
                    CustomsLine("ZERO", "Zero", "PA", 0, Decimal("0.200"), Decimal("4.00")),
                ],
            ),
            CustomsParcel(
                name="Colis 2",
                lines=[CustomsLine("G05-3", "Top", "SH", 6, Decimal("0.200"), Decimal("4.20"))],
            ),
        ],
    )
    path = tmp_path / "liste_colisage.xlsx"

    export_packing_list_xlsx(draft, path)

    workbook = load_workbook(path)
    assert "Résumé" in workbook.sheetnames
    assert "Colisage" in workbook.sheetnames
    summary = workbook["Résumé"]
    sheet = workbook["Colisage"]
    assert summary["B1"].value == "Microstore"
    assert summary["B2"].value == "1001"
    rows = list(sheet.iter_rows(values_only=True))
    assert ("Colis 1", "LA57-5", "Robe", 12, 0.3, 3.6, 5.5, 66) in rows
    assert ("Colis 1", "TOTAL", None, None, None, 3.6, None, 66) in rows
    assert all("ZERO" not in row for row in rows)


def test_customs_declaration_persists_in_database(tmp_path):
    db = Database(tmp_path / "app.sqlite")
    line = InvoiceLine(ref="FL395-2", sage_code="RO", description="Robe", quantity_pieces=12, unit_price_ht=Decimal("5.50"))
    draft = build_customs_draft("Microstore", "1001", [line], {}, {})

    db.save_customs_declaration(draft)
    db.close()

    reopened = Database(tmp_path / "app.sqlite")
    saved = reopened.get_customs_declaration("Microstore", "1001")

    assert saved is not None
    assert saved.parcels[0].lines[0].ref == "FL395-2"
    assert saved.parcels[0].total_weight_kg == Decimal("3.600")
    assert saved.parcels[0].total_value_ht == Decimal("66.00")
    reopened.close()


def test_laposte_script_fills_but_does_not_add_to_cart():
    line = InvoiceLine(ref="FL395-2", sage_code="RO", description="Robe", quantity_pieces=12, unit_price_ht=Decimal("5.50"))
    draft = build_customs_draft("Microstore", "1001", [line], {}, {})

    script = build_laposte_script(draft.parcels[0], draft.parcel_content)

    assert "Commencer votre declaration" in script
    assert "Declarer un objet" in script
    assert "Enregistrer cet objet" in script
    assert "Ajouter au panier" not in script
    assert "62044300" in script


def test_laposte_payload_contains_current_parcel_items():
    line = InvoiceLine(ref="FL395-2", sage_code="RO", description="Robe", quantity_pieces=12, unit_price_ht=Decimal("5.50"))
    draft = build_customs_draft("Microstore", "1001", [line], {}, {})

    payload = parcel_to_laposte_payload(draft.parcels[0], draft.parcel_content)

    assert payload["parcelContent"] == "envoi-commercial"
    assert payload["items"][0]["description"] == "Robe"
    assert payload["items"][0]["unitWeight"] == "0.300"
    assert payload["items"][0]["quantity"] == 12


def test_tampermonkey_userscript_prompts_and_never_adds_to_cart():
    script = (Path(__file__).resolve().parents[1] / "automation" / "laposte_customs.user.js").read_text(encoding="utf-8")

    assert "Remplir depuis Sage Assistant" in script
    assert "left: 18px" in script
    assert "right: 18px" not in script
    assert "localStorage" in script
    assert "confirm(" in script
    assert "GM_xmlhttpRequest" in script
    assert "Ajouter au panier" not in script
