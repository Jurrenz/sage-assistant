from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.db import Database
from app.models import Product
from telegram_bot.commands.stock import HELP_TEXT, authorize, handle_chatid_text, handle_ping_text, handle_stock_text, handle_syncstock_text
from telegram_bot.config import load_bot_settings
from telegram_bot.database.warehouse_repository import WarehouseRepository
from telegram_bot.services.auth_service import AuthService
from telegram_bot.services.stock_import_service import (
    BOX_COUNT_COLUMN,
    PIECES_PER_BOX_COLUMN,
    REF_COLUMN,
    SHEET_NAME,
    TAIL_PIECES_COLUMN,
    StockImportService,
    calculate_total_pieces,
    format_stock_display,
    parse_stock_int,
)
from telegram_bot.services.stock_service import StockService


def make_product(ref: str, package_size: int = 12) -> Product:
    return Product(
        id=None,
        ref=ref,
        type_label="ROBES",
        name="",
        unit_price_ht=Decimal("1.00"),
        package_size=package_size,
    )


def write_stock_file(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.cell(1, REF_COLUMN, "货号")
    sheet.cell(1, TAIL_PIECES_COLUMN, "尾箱")
    sheet.cell(1, PIECES_PER_BOX_COLUMN, "件/箱")
    sheet.cell(1, BOX_COUNT_COLUMN, "箱数")

    sheet.cell(2, REF_COLUMN, "CM217-1")
    sheet.cell(2, TAIL_PIECES_COLUMN, 272)
    sheet.cell(2, PIECES_PER_BOX_COLUMN, 240)
    sheet.cell(2, BOX_COUNT_COLUMN, 1)

    sheet.cell(3, REF_COLUMN, "LA52-7")
    sheet.cell(3, PIECES_PER_BOX_COLUMN, 240)
    sheet.cell(3, BOX_COUNT_COLUMN, 1)

    sheet.cell(4, REF_COLUMN, "UNKNOWN")
    sheet.cell(4, PIECES_PER_BOX_COLUMN, 100)
    sheet.cell(4, BOX_COUNT_COLUMN, 2)

    workbook.save(path)


def test_format_stock_display_and_total_pieces():
    assert format_stock_display(272, 240, 1) == "(272p)+240p×1"
    assert format_stock_display(None, 240, 1) == "240p×1"
    assert calculate_total_pieces(272, 240, 1) == 512
    assert parse_stock_int("(200p)") == 200
    assert parse_stock_int("160p") == 160
    assert parse_stock_int("×2") == 2


def test_stock_import_persists_known_refs_and_ignores_unknown_refs(tmp_path):
    db = Database(tmp_path / "app.sqlite")
    db.upsert_products([make_product("CM217-1"), make_product("LA52-7")])
    stock_path = tmp_path / "stock.xlsx"
    write_stock_file(stock_path)

    repository = WarehouseRepository(db)
    summary = StockImportService(repository).sync_from_xlsx(stock_path)

    assert summary.imported == 2
    assert summary.ignored == 1
    assert summary.ignored_refs == ("UNKNOWN",)

    cm217 = repository.get_stock("CM217-1")
    assert cm217 is not None
    assert cm217.display_text == "(272p)+240p×1"
    assert cm217.total_pieces == 512
    assert cm217.total_packages == 42

    la52 = repository.get_stock("LA52-7")
    assert la52 is not None
    assert la52.display_text == "240p×1"
    assert la52.total_pieces == 240
    assert la52.total_packages == 20

    history_count = db.conn.execute("SELECT COUNT(*) FROM warehouse_stock_history").fetchone()[0]
    assert history_count == 2
    db.close()


def test_stock_service_responses(tmp_path):
    db = Database(tmp_path / "app.sqlite")
    db.upsert_products([make_product("CM217-1"), make_product("NO-STOCK")])
    stock_path = tmp_path / "stock.xlsx"
    write_stock_file(stock_path)

    repository = WarehouseRepository(db)
    StockImportService(repository).sync_from_xlsx(stock_path)
    service = StockService(repository, timezone="Europe/Paris")

    response = service.format_lookup_response("CM217-1")
    assert "Référence : CM217-1" in response
    assert "Stock entrepôt : (272p)+240p×1" in response
    assert "Total : 512 pièces / 42 paquets" in response

    assert service.format_lookup_response("UNKNOWN") == "Référence introuvable."
    assert "stock entrepôt non renseigné" in service.format_lookup_response("NO-STOCK")
    db.close()


def test_command_helpers_authorize_and_sync(tmp_path):
    db = Database(tmp_path / "app.sqlite")
    db.upsert_products([make_product("CM217-1")])
    stock_path = tmp_path / "stock.xlsx"
    write_stock_file(stock_path)

    repository = WarehouseRepository(db)
    stock_service = StockService(repository)
    import_service = StockImportService(repository)

    assert authorize(123, AuthService(set())) is None
    assert authorize(123, AuthService({123})) is None
    denied = authorize(456, AuthService({123}))
    assert denied is not None
    assert denied.text == "Accès refusé."

    sync_response = handle_syncstock_text(import_service, stock_path)
    assert "Préremplissage Excel terminé" in sync_response
    assert "1 références importées" in sync_response
    assert "2 ignorées" in sync_response

    lookup_response = handle_stock_text("/stock CM217-1", stock_service)
    assert "Stock entrepôt : (272p)+240p×1" in lookup_response
    db.close()


def test_config_file_is_created_when_missing(tmp_path, monkeypatch):
    monkeypatch.delenv("SAGE_ASSISTANT_TELEGRAM_TOKEN", raising=False)
    monkeypatch.delenv("SAGE_ASSISTANT_TELEGRAM_CHAT_IDS", raising=False)
    settings_path = tmp_path / "telegram_bot_settings.json"

    settings = load_bot_settings(settings_path)

    assert settings_path.exists()
    assert settings.bot_token == ""
    assert settings.allowed_chat_ids == set()
    assert "allowed_chat_ids" in settings_path.read_text(encoding="utf-8")


def test_config_loads_json_and_environment_override(tmp_path, monkeypatch):
    settings_path = tmp_path / "telegram_bot_settings.json"
    settings_path.write_text(
        '{"bot_token": "json-token", "allowed_chat_ids": [123], "stock_import_path": "stock.xlsx", "timezone": "Europe/Paris"}',
        encoding="utf-8",
    )
    monkeypatch.setenv("SAGE_ASSISTANT_TELEGRAM_TOKEN", "env-token")

    settings = load_bot_settings(settings_path)

    assert settings.bot_token == "env-token"
    assert settings.allowed_chat_ids == {123}


def test_chatid_ping_and_help_text(tmp_path):
    db = Database(tmp_path / "app.sqlite")
    db.upsert_products([make_product("CM217-1")])
    stock_path = tmp_path / "stock.xlsx"
    write_stock_file(stock_path)
    repository = WarehouseRepository(db)
    StockImportService(repository).sync_from_xlsx(stock_path)

    chatid_response = handle_chatid_text(-1001, "supergroup", "Equipe stock", 42)
    assert "Chat ID : -1001" in chatid_response
    assert "Type : supergroup" in chatid_response
    assert "Titre : Equipe stock" in chatid_response
    assert "User ID : 42" in chatid_response

    ping_response = handle_ping_text(db, repository, str(stock_path))
    assert "Bot en ligne" in ping_response
    assert "Base SQLite OK" in ping_response
    assert "Produits : 1" in ping_response
    assert "Stocks entrepôt : 1" in ping_response
    assert "Dernière synchro : " in ping_response

    assert "/chatid" in HELP_TEXT
    assert "/ping" in HELP_TEXT
    assert "warehouse_stock" in HELP_TEXT
    db.close()


def test_real_stock_xlsx_cm217_when_available(tmp_path):
    stock_path = Path(__file__).resolve().parents[2] / "stock.xlsx"
    if not stock_path.exists():
        pytest.skip("stock.xlsx absent du workspace")

    db = Database(tmp_path / "app.sqlite")
    db.upsert_products([make_product("CM217-1")])
    repository = WarehouseRepository(db)

    summary = StockImportService(repository).sync_from_xlsx(stock_path)
    stock = repository.get_stock("CM217-1")

    assert summary.imported >= 1
    assert stock is not None
    assert stock.display_text == "(272p)+240p×1"
    assert stock.total_pieces == 512
    assert stock.total_packages == 42
    db.close()
